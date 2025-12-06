import sys
import os
import pandas as pd
import openpyxl
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QPushButton, QLabel, QFileDialog, QMessageBox, QFrame)
from PyQt6.QtCore import Qt


class ExcelMinimalApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Excel Merger Pro")
        self.setGeometry(100, 100, 400, 350)

        # --- НАСТРОЙКИ (КОНСТАНТЫ) ---
        # Здесь мы запоминаем названия колонок навсегда
        self.KEY_COLUMN = "Артикул"
        self.COL_1 = 'Продажи ("Оплата за клики"), ₽'
        self.COL_2 = 'Продажи ("Оплата за заказ"), ₽'
        # -----------------------------

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)
        self.layout.setSpacing(15)  # Отступы между элементами

        self.file_main_path = None
        self.file_donor_path = None

        # === ИНТЕРФЕЙС ===

        # 1. Главный файл
        self.lbl_step1 = QLabel("1. Выберите файл, В КОТОРЫЙ добавляем:")
        self.lbl_step1.setStyleSheet("font-weight: bold; font-size: 14px;")
        self.layout.addWidget(self.lbl_step1)

        self.btn_main = QPushButton("📂 Выбрать главный файл")
        self.btn_main.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_main.clicked.connect(self.select_main_file)
        self.layout.addWidget(self.btn_main)

        self.lbl_main_status = QLabel("Файл не выбран")
        self.lbl_main_status.setStyleSheet("color: #666; font-size: 11px; margin-bottom: 5px;")
        self.layout.addWidget(self.lbl_main_status)

        # Разделитель
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        self.layout.addWidget(line)

        # 2. Файл донор
        self.lbl_step2 = QLabel("2. Выберите файл, ИЗ КОТОРОГО берем:")
        self.lbl_step2.setStyleSheet("font-weight: bold; font-size: 14px;")
        self.layout.addWidget(self.lbl_step2)

        self.btn_donor = QPushButton("📂 Выбрать файл с продажами")
        self.btn_donor.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_donor.clicked.connect(self.select_donor_file)
        self.layout.addWidget(self.btn_donor)

        self.lbl_donor_status = QLabel("Файл не выбран")
        self.lbl_donor_status.setStyleSheet("color: #666; font-size: 11px; margin-bottom: 5px;")
        self.layout.addWidget(self.lbl_donor_status)

        self.layout.addStretch()  # Сдвигаем кнопку запуска вниз

        # 3. Кнопка запуска
        self.btn_run = QPushButton("ЗАПУСТИТЬ И СОХРАНИТЬ")
        self.btn_run.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32; 
                color: white; 
                font-weight: bold; 
                font-size: 14px; 
                padding: 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
        """)
        self.btn_run.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_run.clicked.connect(self.process_files)
        self.layout.addWidget(self.btn_run)

    def select_main_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Выбрать главный файл", "", "Excel Files (*.xlsx *.xls)")
        if file:
            self.file_main_path = file
            self.lbl_main_path_short = os.path.basename(file)  # Показываем только имя файла
            self.lbl_main_status.setText(f"✅ {self.lbl_main_path_short}")
            self.lbl_main_status.setStyleSheet("color: green; font-size: 11px;")

    def select_donor_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Выбрать файл-донор", "", "Excel Files (*.xlsx *.xls)")
        if file:
            self.file_donor_path = file
            self.lbl_donor_path_short = os.path.basename(file)
            self.lbl_donor_status.setText(f"✅ {self.lbl_donor_path_short}")
            self.lbl_donor_status.setStyleSheet("color: green; font-size: 11px;")

    def find_header_row(self, filepath, key_column):
        """Автопоиск строки заголовка"""
        try:
            # Сканируем первые 50 строк
            df_temp = pd.read_excel(filepath, header=None, nrows=50)
            for idx, row in df_temp.iterrows():
                row_values = [str(x).strip() for x in row.values]
                if key_column in row_values:
                    return idx
            return None
        except Exception:
            return None

    def process_files(self):
        # 1. Проверка, выбраны ли файлы
        if not self.file_main_path or not self.file_donor_path:
            QMessageBox.warning(self, "Внимание", "Пожалуйста, выберите оба файла!")
            return

        # 2. Сразу спрашиваем, куда сохранить результат (Save As Dialog)
        default_name = "Результат_сборки.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить результат как...", default_name,
                                                   "Excel Files (*.xlsx)")

        if not save_path:
            return  # Пользователь передумал и закрыл окно сохранения

        try:
            # 3. Автопоиск заголовков
            donor_header_row = self.find_header_row(self.file_donor_path, self.KEY_COLUMN)
            if donor_header_row is None:
                QMessageBox.critical(self, "Ошибка структуры",
                                     f"Не могу найти колонку '{self.KEY_COLUMN}' в файле-доноре.\nПроверьте, что это верный файл.")
                return

            main_header_row = self.find_header_row(self.file_main_path, self.KEY_COLUMN)
            if main_header_row is None:
                main_header_row = 0

            # 4. Читаем данные (Pandas)
            df_main = pd.read_excel(self.file_main_path, header=main_header_row)
            df_donor = pd.read_excel(self.file_donor_path, header=donor_header_row)

            # Проверяем, есть ли нужные колонки в доноре
            missing = []
            if self.COL_1 not in df_donor.columns: missing.append(self.COL_1)
            if self.COL_2 not in df_donor.columns: missing.append(self.COL_2)

            if missing:
                QMessageBox.critical(self, "Нет колонок",
                                     f"В файле-доноре не найдены колонки:\n{', '.join(missing)}")
                return

            # Чистим Артикулы (убираем пробелы, делаем строкой)
            df_main[self.KEY_COLUMN] = df_main[self.KEY_COLUMN].astype(str).str.strip()
            df_donor[self.KEY_COLUMN] = df_donor[self.KEY_COLUMN].astype(str).str.strip()

            # Убираем дубликаты артикулов в доноре
            donor_clean = df_donor[[self.KEY_COLUMN, self.COL_1, self.COL_2]].drop_duplicates(subset=[self.KEY_COLUMN])

            # Объединяем (VLOOKUP)
            merged = pd.merge(df_main, donor_clean, on=self.KEY_COLUMN, how='left')
            merged[self.COL_1] = merged[self.COL_1].fillna(0)
            merged[self.COL_2] = merged[self.COL_2].fillna(0)

            # 5. Записываем в файл с сохранением формата (OpenPyXL)
            wb = openpyxl.load_workbook(self.file_main_path)
            ws = wb.active

            excel_header_row = main_header_row + 1
            max_col = ws.max_column

            # Ищем индексы колонок (куда писать)
            col1_idx = None
            col2_idx = None

            for col in range(1, max_col + 1):
                val = ws.cell(row=excel_header_row, column=col).value
                if val == self.COL_1: col1_idx = col
                if val == self.COL_2: col2_idx = col

            # Создаем колонки, если их нет
            if col1_idx is None:
                col1_idx = max_col + 1
                ws.cell(row=excel_header_row, column=col1_idx).value = self.COL_1
                max_col += 1

            if col2_idx is None:
                col2_idx = max_col + 1
                ws.cell(row=excel_header_row, column=col2_idx).value = self.COL_2

            # Записываем данные построчно
            for i in range(len(merged)):
                val1 = merged.iloc[i][self.COL_1]
                val2 = merged.iloc[i][self.COL_2]

                row_idx = i + excel_header_row + 1

                ws.cell(row=row_idx, column=col1_idx).value = val1
                ws.cell(row=row_idx, column=col2_idx).value = val2

            # Сохраняем по новому пути
            wb.save(save_path)

            QMessageBox.information(self, "Успех", f"Готово! Файл сохранен:\n{save_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка:\n{str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelMinimalApp()
    window.show()
    sys.exit(app.exec())